from io import BytesIO
from urllib.parse import quote

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.text import slugify

from accounts.models import User

from .forms import ReportUserForm
from .models import HomePage, Match, Swipe, UserBan, UserBlock, UserRecommendation
from .services import record_swipe


FEED_FACT_SECTIONS = [
    {
        "id": "communication",
        "title": "Коммуникация",
        "items": [
            "Лучше всего работают вопросы, которые нельзя ответить одним словом.",
            "Не начинай с оценки внешности — спроси про интерес или ценность.",
            "Короткое сообщение + конкретика = выше шанс на ответ.",
        ],
    },
    {
        "id": "family",
        "title": "Семья и отношения",
        "items": [
            "Совместимость чаще зависит от привычек и границ, чем от хобби.",
            "Обсуждение ожиданий на раннем этапе экономит время обоим.",
            "Уважение к личному времени — один из самых сильных маркеров зрелости.",
        ],
    },
    {
        "id": "safety",
        "title": "Безопасность",
        "items": [
            "Первые встречи — в людных местах и с понятным планом.",
            "Не отправляй документы/пароли/коды подтверждения — никому и никогда.",
            "Если что-то смущает — лучше отложить общение и спросить напрямую.",
        ],
    },
    {
        "id": "profile",
        "title": "Профиль",
        "items": [
            "1-2 чёткие фото лица + 1 фото в полный рост делают профиль понятнее.",
            "Короткая био: кто ты, что ценишь, что ищешь — без общих фраз.",
            "Анкета помогает рекомендации стать точнее: заполняй постепенно.",
        ],
    },
]


def landing(request):
    if request.user.is_authenticated:
        return redirect("feed")

    page = HomePage.objects.filter(slug="landing", is_active=True).first()
    blocks = []
    if page is not None:
        blocks = list(page.blocks.filter(is_active=True).order_by("order", "id"))

    return render(request, "matchmaking/landing.html", {"cms_blocks": blocks})


def _recommendation_pair_or_404(request, user_id: int):
    other_user = get_object_or_404(User, id=user_id)
    if other_user.id == request.user.id:
        raise Http404

    other_profile = getattr(other_user, "profile", None)
    my_profile = getattr(request.user, "profile", None)
    if other_profile is None or my_profile is None:
        raise Http404

    rec = (
        UserRecommendation.objects.filter(to_user=request.user, recommended_user=other_user)
        .order_by("-created_at", "-id")
        .first()
    )
    if rec is None:
        raise Http404

    return rec, my_profile, other_user, other_profile


@login_required
def recommendation_compatibility(request, user_id: int):
    rec, my_profile, other_user, other_profile = _recommendation_pair_or_404(request, user_id)

    from .compatibility import compatibility_breakdown

    report = compatibility_breakdown(my_profile, other_profile)
    sections = report.get("sections") or []

    def _label_value(value, choices_map: dict):
        if value is None:
            return "—"
        if isinstance(value, (list, tuple, set)):
            parts = []
            for item in value:
                item_s = str(item)
                parts.append(str(choices_map.get(item_s, item_s)))
            return ", ".join(parts) if parts else "—"
        value_s = str(value)
        return str(choices_map.get(value_s, value_s))

    for section in sections:
        for q in section.get("questions") or []:
            choices_map = {str(v): str(lbl) for v, lbl in (q.get("choices") or [])}
            for key in ("a_to_b", "b_to_a"):
                part = q.get(key)
                if not part:
                    continue
                part["score_percent"] = int(round(float(part.get("score") or 0.0) * 100))
                part["expected_label"] = _label_value(part.get("expected"), choices_map)
                part["actual_label"] = _label_value(part.get("actual"), choices_map)

    chart_labels = [s.get("title") or s.get("id") or "" for s in sections]
    chart_values_a_to_b = [s.get("a_to_b") for s in sections]
    chart_values_b_to_a = [s.get("b_to_a") for s in sections]

    return render(
        request,
        "matchmaking/recommendation_compatibility.html",
        {
            "recommendation": rec,
            "other_user": other_user,
            "other_profile": other_profile,
            "report": report,
            "sections": sections,
            "chart_labels": chart_labels,
            "chart_values_a_to_b": chart_values_a_to_b,
            "chart_values_b_to_a": chart_values_b_to_a,
        },
    )


@login_required
def recommendation_excel(request, user_id: int):
    rec, my_profile, other_user, other_profile = _recommendation_pair_or_404(request, user_id)

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    from profiles.questionnaire import get_questionnaire_spec_for_profile

    wb = Workbook()
    wb.remove(wb.active)

    def _label_value(value, choices_map: dict[str, str]):
        if value is None:
            return ""
        if isinstance(value, (list, tuple, set)):
            out = []
            for item in value:
                item_s = str(item)
                out.append(choices_map.get(item_s, item_s))
            return ", ".join([x for x in out if str(x).strip() != ""])
        value_s = str(value)
        return choices_map.get(value_s, value_s)

    def add_sheet(profile, kind: str, title: str):
        spec = get_questionnaire_spec_for_profile(profile, kind)
        answers = profile.questionnaire_me if kind == "me" else profile.questionnaire_ideal
        answers = answers or {}

        ws = wb.create_sheet(title=title)
        ws.append(["Раздел", "Вопрос", "Ответ (значение)", "Ответ (текст)"])

        for section in spec:
            section_title = section.get("title") or section.get("id") or ""
            for q in section.get("questions") or []:
                qid = q.get("id")
                if not qid:
                    continue
                q_text = q.get("text") or ""

                raw = answers.get(qid)
                choices_map = {str(v): str(lbl) for v, lbl in (q.get("choices") or [])}

                if isinstance(raw, (list, tuple, set)):
                    raw_value = ", ".join([str(x) for x in raw])
                else:
                    raw_value = "" if raw is None else str(raw)

                ws.append([section_title, q_text, raw_value, _label_value(raw, choices_map)])

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 32 if col in (1, 2) else 24

        ws.freeze_panes = "A2"

    add_sheet(my_profile, "me", "Мои ответы")
    add_sheet(my_profile, "ideal", "Мой идеал")

    meta = wb.create_sheet(title="Инфо")
    meta.append(["Рекомендация ID", rec.id])
    meta.append(["Кому отправлено", request.user.username])
    meta.append(["Кандидат", other_user.username])
    meta.append(["Скор", rec.score if rec.score is not None else ""]) 
    meta.append(["Создана", rec.created_at.strftime("%Y-%m-%d %H:%M") if rec.created_at else ""]) 

    base = slugify(other_user.username) or "candidate"
    filename = f"recommendation_{base}_{rec.id}.xlsx"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"; filename*=UTF-8\'\'{quote(filename)}'
    )
    return response


@login_required
def block_user(request, user_id: int):
    if request.method != "POST":
        raise Http404

    to_user = get_object_or_404(User, id=user_id)
    if to_user.id == request.user.id:
        raise Http404

    UserBlock.objects.get_or_create(blocker=request.user, blocked=to_user)

    rec_qs = UserRecommendation.objects.filter(
        to_user=request.user,
        recommended_user=to_user,
        consumed_at__isnull=True,
    )
    now = timezone.now()
    rec_qs.filter(seen_at__isnull=True).update(seen_at=now)
    rec_qs.update(consumed_at=now)

    Match.objects.filter(
        Q(user1=request.user, user2=to_user) | Q(user1=to_user, user2=request.user)
    ).delete()
    Swipe.objects.filter(
        Q(from_user=request.user, to_user=to_user) | Q(from_user=to_user, to_user=request.user)
    ).delete()

    messages.success(request, "Пользователь заблокирован.")

    if request.headers.get("HX-Request") == "true":
        candidate, recommendation = _next_candidate_for(request.user)
        ctx = {"candidate": candidate, "recommendation": recommendation}
        return render(request, "matchmaking/_card.html", ctx)

    next_url = request.POST.get("next")
    if next_url and str(next_url).startswith("/"):
        return redirect(next_url)
    return redirect("feed")


@login_required
def unblock_user(request, user_id: int):
    if request.method != "POST":
        raise Http404

    to_user = get_object_or_404(User, id=user_id)
    if to_user.id == request.user.id:
        raise Http404

    UserBlock.objects.filter(blocker=request.user, blocked=to_user).delete()
    messages.success(request, "Блокировка снята.")

    next_url = request.POST.get("next")
    if next_url and str(next_url).startswith("/"):
        return redirect(next_url)
    return redirect("me")


@login_required
def report_user(request, user_id: int):
    target = get_object_or_404(User, id=user_id)
    if target.id == request.user.id:
        raise Http404

    if request.method == "POST":
        form = ReportUserForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.reporter = request.user
            report.reported_user = target
            report.save()
            messages.success(request, "Жалоба отправлена. Спасибо!")
            return redirect("public_profile", user_id=target.id)
    else:
        form = ReportUserForm()

    target_profile = target.profile if hasattr(target, "profile") else None
    target_name = None
    if target_profile is not None:
        target_name = target_profile.display_name
    if not target_name:
        target_name = target.get_username()

    return render(
        request,
        "matchmaking/report_user.html",
        {
            "form": form,
            "target": target,
            "target_name": target_name,
        },
    )

@login_required
def undo_swipe(request):
    if request.method != "POST":
        raise Http404

    last = Swipe.objects.filter(from_user=request.user).order_by("-created_at").first()
    if last is not None:
        rec = (
            UserRecommendation.objects.filter(
                to_user=request.user,
                recommended_user=last.to_user,
                consumed_at__isnull=False,
            )
            .order_by("-consumed_at", "-id")
            .first()
        )
        if rec is not None:
            rec.consumed_at = None
            rec.save(update_fields=["consumed_at"])
        last.delete()

    candidate, recommendation = _next_candidate_for(request.user)
    ctx = {"candidate": candidate, "recommendation": recommendation}

    if request.headers.get("HX-Request") == "true":
        return render(request, "matchmaking/_card.html", ctx)

    return redirect("feed")


def _next_candidate_for(user):
    banned_ids = UserBan.objects.active().values_list("user_id", flat=True)
    blocked_ids = UserBlock.objects.filter(blocker=user).values_list("blocked_id", flat=True)
    blocked_by_ids = UserBlock.objects.filter(blocked=user).values_list("blocker_id", flat=True)

    rec = (
        UserRecommendation.objects.filter(to_user=user, consumed_at__isnull=True)
        .filter(recommended_user__is_active=True)
        .filter(recommended_user__profile__isnull=False)
        .exclude(recommended_user_id__in=banned_ids)
        .exclude(recommended_user_id__in=blocked_ids)
        .exclude(recommended_user_id__in=blocked_by_ids)
        .select_related("recommended_user", "recommended_user__profile")
        .order_by("-created_at", "-id")
        .first()
    )

    if rec is None:
        return None, None

    if rec.seen_at is None:
        rec.seen_at = timezone.now()
        rec.save(update_fields=["seen_at"])

    return rec.recommended_user.profile, rec


@login_required
def feed(request):
    candidate, recommendation = _next_candidate_for(request.user)
    ctx = {
        "candidate": candidate,
        "recommendation": recommendation,
        "fact_sections": FEED_FACT_SECTIONS,
    }

    if request.headers.get("HX-Request") == "true":
        return render(request, "matchmaking/_card.html", ctx)

    return render(request, "matchmaking/feed.html", ctx)


@login_required
def swipe(request, user_id: int, value: str):
    if request.method != "POST":
        raise Http404

    if value not in (Swipe.Value.LIKE, Swipe.Value.PASS):
        raise Http404

    to_user = get_object_or_404(User, id=user_id)

    if to_user.id == request.user.id:
        raise Http404

    if not to_user.is_active:
        raise Http404

    if UserBan.objects.active().filter(user=to_user).exists():
        raise Http404

    if UserBlock.objects.filter(
        Q(blocker=request.user, blocked=to_user) | Q(blocker=to_user, blocked=request.user)
    ).exists():
        raise Http404

    _, created_match = record_swipe(from_user=request.user, to_user=to_user, value=value)

    rec_qs = UserRecommendation.objects.filter(
        to_user=request.user,
        recommended_user=to_user,
        consumed_at__isnull=True,
    )
    now = timezone.now()
    rec_qs.filter(seen_at__isnull=True).update(seen_at=now)
    rec_qs.update(consumed_at=now)
    other_profile = getattr(to_user, "profile", None)
    other_name = None
    if other_profile is not None:
        other_name = other_profile.display_name
    if not other_name:
        other_name = to_user.get_username()

    if created_match is not None:
        messages.success(request, f"Мэтч! Теперь вы можете написать {other_name}.")

    candidate, recommendation = _next_candidate_for(request.user)
    ctx = {"candidate": candidate, "recommendation": recommendation}

    if created_match is not None and request.headers.get("HX-Request") == "true":
        ctx["match_name"] = other_name

    if request.headers.get("HX-Request") == "true":
        return render(request, "matchmaking/_card.html", ctx)

    return redirect("feed")


@login_required
def matches(request):
    blocked_ids = set(
        UserBlock.objects.filter(blocker=request.user).values_list("blocked_id", flat=True)
    )
    blocked_by_ids = set(
        UserBlock.objects.filter(blocked=request.user).values_list("blocker_id", flat=True)
    )

    qs = (
        Match.objects.filter(Q(user1=request.user) | Q(user2=request.user))
        .select_related(
            "user1",
            "user2",
            "user1__profile",
            "user2__profile",
        )
        .prefetch_related("messages")
        .order_by("-created_at")
    )

    banned_ids = set(UserBan.objects.active().values_list("user_id", flat=True))

    rows = []
    for m in qs:
        other = m.other(request.user)
        if not other.is_active or other.id in banned_ids:
            continue
        if other.id in blocked_ids or other.id in blocked_by_ids:
            continue

        other_profile = other.profile if hasattr(other, "profile") else None
        other_name = None
        if other_profile is not None:
            other_name = other_profile.display_name
        if not other_name:
            other_name = other.get_username()

        other_avatar_url = None
        if other_profile is not None and other_profile.avatar:
            other_avatar_url = other_profile.avatar.url

        # Get last message preview
        last_message = m.messages.order_by("-created_at").first()
        last_message_text = ""
        last_message_time = None
        if last_message:
            last_message_text = last_message.text[:100]
            last_message_time = last_message.created_at

        rows.append(
            {
                "match": m,
                "other": other,
                "other_name": other_name,
                "other_avatar_url": other_avatar_url,
                "last_message_text": last_message_text,
                "last_message_time": last_message_time,
            }
        )
    return render(request, "matchmaking/matches.html", {"rows": rows})
