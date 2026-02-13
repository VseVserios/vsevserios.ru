from __future__ import annotations

from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


SECTION_CODE_OVERRIDES = {
    "Жизненные принципы": "principles",
    "Жилищные условия": "housing",
    "Роли": "roles",
    "Работа": "work",
    "Расходы на жену": "wife_expenses",
    "Сексуальная совместимость": "sexual",
    "Отдых": "rest",
    "Язык любви": "love_language",
    "Тест": "tests",
    "Тесты": "tests",
}


def _parse_bool(value) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    s = str(value).strip().lower()
    return s in {"1", "true", "yes", "y", "да", "+", "x"}


def _guess_is_multiple(question_text: str) -> bool:
    t = (question_text or "").strip().lower()
    if not t:
        return False
    if "можно выбрать несколько" in t:
        return True
    if "выберите несколько" in t:
        return True
    if "несколько вариантов" in t:
        return True
    if "несколько ответ" in t:
        return True
    if "отмет" in t and "несколь" in t:
        return True
    return False


def _split_choices(value) -> list[str]:
    if value is None:
        return []
    raw = str(value)
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")

    for sep in (";", "\n", "|"):
        if sep in raw:
            parts = [p.strip() for p in raw.split(sep)]
            parts = [p for p in parts if p]
            if parts:
                return parts

    if "," in raw:
        raw = raw.replace("\n", " ")
        parts = [p.strip() for p in raw.split(",")]
        return [p for p in parts if p]

    raw = raw.replace("\n", " ").strip()
    return [raw] if raw else []


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _is_headerish_text(value: str) -> bool:
    t = (value or "").strip().lower()
    if not t:
        return True
    return t in {"вопрос", "вопросы", "ответ", "ответы"}


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _heuristic_find_columns(ws):
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row <= 0 or max_col <= 0:
        return None

    scan_rows = min(max_row, 250)
    scan_cols = min(max_col, 80)

    question_scores = []
    for col in range(1, scan_cols + 1):
        score = 0
        for row in range(1, scan_rows + 1):
            t = _cell_text(ws.cell(row=row, column=col).value)
            if not t or _is_headerish_text(t):
                continue
            if len(t) >= 12:
                score += 1
        question_scores.append((score, col))

    question_scores.sort(reverse=True)
    if not question_scores or question_scores[0][0] == 0:
        return None

    question_col = question_scores[0][1]

    answer_scores = []
    for col in range(1, scan_cols + 1):
        if col == question_col:
            continue
        score = 0
        for row in range(1, scan_rows + 1):
            t = _cell_text(ws.cell(row=row, column=col).value)
            if not t or _is_headerish_text(t):
                continue
            if len(_split_choices(t)) >= 2:
                score += 1
        answer_scores.append((score, col))

    answer_scores.sort(reverse=True)
    answers_col = answer_scores[0][1] if answer_scores and answer_scores[0][0] > 0 else None

    for row in range(1, max_row + 1):
        q_text = _cell_text(ws.cell(row=row, column=question_col).value)
        if not q_text or _is_headerish_text(q_text):
            continue

        if answers_col is not None:
            a_text = _cell_text(ws.cell(row=row, column=answers_col).value)
            if len(_split_choices(a_text)) >= 2:
                return max(0, row - 1), question_col, answers_col, None

        row_choices = []
        for c in range(1, scan_cols + 1):
            if c == question_col:
                continue
            t = _cell_text(ws.cell(row=row, column=c).value)
            if not t or _is_headerish_text(t):
                continue
            row_choices.append(t)
        if len(row_choices) >= 2:
            return max(0, row - 1), question_col, None, None

    return None


def _find_header(ws):
    max_scan = min(ws.max_row or 0, 20)
    for row_idx in range(1, max_scan + 1):
        headers = [_normalize_header(ws.cell(row=row_idx, column=c).value) for c in range(1, ws.max_column + 1)]
        question_col = None
        answers_col = None
        multiple_col = None

        for col_idx, h in enumerate(headers, start=1):
            if not h:
                continue
            if question_col is None and "вопрос" in h:
                question_col = col_idx
            if answers_col is None and ("ответы" in h or h == "ответ"):
                answers_col = col_idx
            if multiple_col is None and any(x in h for x in ("multiple", "муль", "несколь", "множ")):
                multiple_col = col_idx

        if question_col is not None and answers_col is not None:
            return row_idx, question_col, answers_col, multiple_col

    return _heuristic_find_columns(ws)


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            default=str(settings.BASE_DIR / "Анкета.xlsx"),
            help="Path to .xlsx file",
        )
        parser.add_argument(
            "--only-tests",
            action="store_true",
            help="Import only sheet 'Тест'/'Тесты' and replace only the corresponding DB section.",
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Actually write changes to DB (dangerous). Without this flag, runs in dry-run mode.",
        )
        parser.add_argument(
            "--all-multiple",
            action="store_true",
            help="Mark all imported questions as multiple-choice.",
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise CommandError("openpyxl is required. Install it and try again.") from e

        file_path = Path(options["path"]).expanduser()
        if not file_path.is_absolute():
            file_path = settings.BASE_DIR / file_path
        if not file_path.exists():
            raise CommandError(f"Excel file not found: {file_path}")

        try:
            wb = load_workbook(filename=str(file_path), data_only=True)
        except PermissionError as e:
            raise CommandError(
                "Cannot open Excel file (permission denied). "
                "Close the file in Excel, wait for OneDrive sync to finish, "
                "or copy it to a local folder (e.g. C:\\Temp) and re-run with --path. "
                f"File: {file_path}"
            ) from e
        except OSError as e:
            raise CommandError(f"Cannot open Excel file: {file_path} ({e})") from e

        only_tests = bool(options.get("only_tests"))

        parsed_sections = []
        used_section_codes: set[str] = set()
        used_question_codes: set[str] = set()

        for sheet_order, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            is_tests_sheet = str(sheet_name).strip().lower() in {"тест", "тесты"}
            if only_tests and not is_tests_sheet:
                if len(wb.sheetnames) != 1:
                    continue
            header = _find_header(ws)
            if header is None:
                raise CommandError(f"Cannot find header row with 'Вопрос'/'Ответы' in sheet: {sheet_name}")

            header_row, question_col, answers_col, multiple_col = header

            if only_tests:
                base_section_code = "tests"
            else:
                base_section_code = SECTION_CODE_OVERRIDES.get(str(sheet_name).strip())
                if not base_section_code:
                    base_section_code = f"section_{sheet_order:02d}"

            section_code = base_section_code
            suffix = 2
            while section_code in used_section_codes:
                section_code = f"{base_section_code}_{suffix}"
                suffix += 1
            used_section_codes.add(section_code)

            questions = []
            question_order = 0

            for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
                q_raw = ws.cell(row=row_idx, column=question_col).value
                if q_raw is None:
                    continue
                q_text = str(q_raw).strip()
                if not q_text:
                    continue

                if answers_col is not None:
                    a_raw = ws.cell(row=row_idx, column=answers_col).value
                    choice_labels = _split_choices(a_raw)
                else:
                    choice_labels = []
                    max_col = ws.max_column or 0
                    for c in range(question_col + 1, max_col + 1):
                        if multiple_col is not None and c == multiple_col:
                            continue
                        t = _cell_text(ws.cell(row=row_idx, column=c).value)
                        if not t or _is_headerish_text(t):
                            continue
                        parts = _split_choices(t)
                        if parts:
                            choice_labels.extend(parts)

                    seen = set()
                    deduped = []
                    for label in choice_labels:
                        key = str(label).strip()
                        if not key or key in seen:
                            continue
                        seen.add(key)
                        deduped.append(key)
                    choice_labels = deduped

                input_type = "choice"
                if not choice_labels:
                    input_type = "text"

                question_order += 1

                q_code = f"{section_code}_{question_order:02d}"
                if q_code in used_question_codes:
                    extra = 2
                    q_code_candidate = f"{q_code}_{extra}"
                    while q_code_candidate in used_question_codes:
                        extra += 1
                        q_code_candidate = f"{q_code}_{extra}"
                    q_code = q_code_candidate
                used_question_codes.add(q_code)

                row_is_multiple = False
                if options.get("all_multiple"):
                    row_is_multiple = True
                elif multiple_col is not None:
                    row_is_multiple = _parse_bool(ws.cell(row=row_idx, column=multiple_col).value)
                else:
                    row_is_multiple = _guess_is_multiple(q_text)

                if input_type == "text":
                    row_is_multiple = False

                questions.append(
                    {
                        "code": q_code,
                        "text": q_text,
                        "input_type": input_type,
                        "is_multiple": row_is_multiple,
                        "choices": choice_labels,
                    }
                )

            if not questions:
                continue

            parsed_sections.append(
                {
                    "code": section_code,
                    "title": "Тесты" if only_tests else str(sheet_name).strip(),
                    "order": sheet_order,
                    "questions": questions,
                }
            )

        section_count = len(parsed_sections)
        question_count = sum(len(s["questions"]) for s in parsed_sections)
        choice_count = sum(len(q["choices"]) for s in parsed_sections for q in s["questions"])

        self.stdout.write(
            f"Parsed: sections={section_count} questions={question_count} choices={choice_count} file={file_path}"
        )

        if not options.get("apply"):
            self.stdout.write("Dry-run mode. Re-run with --apply to write to DB.")
            return

        from profiles.models import QuestionnaireChoice, QuestionnaireQuestion, QuestionnaireSection

        section_field_names = {f.name for f in QuestionnaireSection._meta.get_fields()}

        with transaction.atomic():
            if only_tests:
                tests_sections = QuestionnaireSection.objects.filter(code="tests")
                tests_questions = QuestionnaireQuestion.objects.filter(section__in=tests_sections)
                QuestionnaireChoice.objects.filter(question__in=tests_questions).delete()
                tests_questions.delete()
                tests_sections.delete()
            else:
                QuestionnaireChoice.objects.all().delete()
                QuestionnaireQuestion.objects.all().delete()
                QuestionnaireSection.objects.all().delete()

            for section in parsed_sections:
                section_kwargs = {
                    "code": section["code"],
                    "gender": "",
                    "title": section["title"],
                    "hint": "",
                    "order": section["order"],
                }

                section_obj = QuestionnaireSection.objects.create(**section_kwargs)

                for q_order, q in enumerate(section["questions"], start=1):
                    question_obj = QuestionnaireQuestion.objects.create(
                        section=section_obj,
                        code=q["code"],
                        gender="",
                        text=q["text"],
                        input_type=str(q.get("input_type") or "choice"),
                        is_multiple=bool(q.get("is_multiple")),
                        order=q_order,
                    )

                    for c_order, label in enumerate(q["choices"], start=1):
                        QuestionnaireChoice.objects.create(
                            question=question_obj,
                            value=str(c_order),
                            label=str(label).strip(),
                            order=c_order,
                        )

        self.stdout.write("Import completed.")
